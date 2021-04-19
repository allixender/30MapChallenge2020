import ogr
import json
import argparse
from openpyxl import load_workbook
import os, sys

# Initialize parser 
parser = argparse.ArgumentParser(description='convert xlsx or shp input to basic json')

parser.add_argument("-o", "--outfile", help = "Output, default data_new.json")
parser.add_argument("-e", "--xlsx-in", help = "excel input file, e.g.: andmed-24032020.xlsx")
parser.add_argument("-s", "--shp-in", help = "shapefile input file, e.g.: rahvastik_maakond.shp")

args = parser.parse_args()  

out_json_file = "data_new.json"
in_shapefile = "rahvastik_maakond_covid_uus2.shp"
in_excelfile = "koroonaviiruse andmed 24.03.2020.xlsx"

def get_excel_data(path_to_excel):
    new_data = {}
    new_data['features'] = []

    wb = load_workbook(filename = path_to_excel, read_only=True)
    print(wb.sheetnames)
    ws = wb['Sheet1']
    for x in range(3,17):
        data_point = {}
        data_point['NIMI'] = ws.cell(row=x, column=1).value
        data_point['covid_arv'] = ws.cell(row=x, column=3).value
        new_data['features'].append(data_point)
    return new_data

def get_shape_data(path_to_shape):
    new_data = {}
    new_data['features'] = []

    in_driver = ogr.GetDriverByName("ESRI Shapefile")
    in_data_source = in_driver.Open(in_shapefile, 0)
    in_layer = in_data_source.GetLayer()
    layer_definition = in_layer.GetLayerDefn()

    # MNIMI,MKOOD,nr,KOOD,FID_1,VID,NIMI,STAMP_CRE,pop_total,60yle,60yle_prop,covid_1000,covid13_03,covid14_03,covid15_03,covid16_03,covid17_03,covid18_03,covid19_03,covid20_03,covid21_03,covid22_03,covid23_03
    for i in range(0, in_layer.GetFeatureCount()):
        # Get the input Feature
        in_feature = in_layer.GetFeature(i)
        data_point = {}
        for i in range(layer_definition.GetFieldCount()):
            field_name = layer_definition.GetFieldDefn(i).GetName()
            data_point[field_name] = in_feature.GetField(field_name)
        
        new_data['features'].append(data_point)
    return new_data


if __name__ == "__main__":

    # print(args)
    # Namespace(
    # outfile='data_new.json', 
    # shp_in='rahvastik_maakond_covid_uus2.shp', 
    # xlsx_in='koroonaviiruse andmed 24.03.2020.xlsx')

    new_data = {}

    if args.shp_in is not None and args.xlsx_in is not None:
        print("please only provide either xlsx or shp input, not both")
        sys.exit(1)
    elif args.shp_in is not None:
        # Get the input Layer
        # in_shapefile = "rahvastik_maakond_covid_uus2.shp"
        in_shapefile = os.path.join(os.getcwd(), args.shp_in)
        new_data = get_shape_data(in_shapefile)
    elif args.xlsx_in is not None:
        in_excelfile = os.path.join(os.getcwd(), args.xlsx_in)
        new_data = get_excel_data(in_excelfile)
    else:
        print("no input options given")
        sys.exit(1)
    
    if args.outfile is not None:
        out_json_file = args.outfile

    with open(out_json_file, 'w') as ojs:
        ojs.write(json.dumps(new_data, indent=4))
